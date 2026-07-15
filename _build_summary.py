# -*- coding: utf-8 -*-
from pathlib import Path
import openpyxl
import json

BASE = Path(r"C:\Users\uriak_wboloix\medusa\group-buying-site")
FILES = list(BASE.glob("\uc544\uc774\ub3cc\uacf5\uad6c-*.xlsx"))
# prefer exact two files without (1) duplicate
screen = BASE / "\uc544\uc774\ub3cc\uacf5\uad6c-\ud654\uba74\uc815\uc758\uc11c.xlsx"
func = BASE / "\uc544\uc774\ub3cc\uacf5\uad6c-\uae30\ub2a5\uc815\uc758\uc11c.xlsx"
OUT = BASE / "_spec_extract.txt"
MAX_ROWS = 100

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

def extract_workbook(path, out_lines):
    out_lines.append("=" * 80)
    out_lines.append("FILE: " + path.name)
    out_lines.append("PATH: " + str(path))
    out_lines.append("=" * 80)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out_lines.append("SHEET_COUNT: " + str(len(wb.sheetnames)))
    out_lines.append("SHEET_NAMES: " + json.dumps(wb.sheetnames, ensure_ascii=False))
    out_lines.append("")
    for sn in wb.sheetnames:
        ws = wb[sn]
        out_lines.append("-" * 60)
        out_lines.append("SHEET: " + sn)
        out_lines.append("max_row=" + str(ws.max_row) + ", max_column=" + str(ws.max_column))
        out_lines.append("-" * 60)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= MAX_ROWS:
                out_lines.append("... truncated after " + str(MAX_ROWS) + " rows ...")
                break
            cells = [cell_str(c) for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if not any(cells):
                continue
            out_lines.append("\t".join(cells))
            row_count += 1
        out_lines.append("(rows written: " + str(row_count) + ")")
        out_lines.append("")
    wb.close()

STOREFRONT = """
STOREFRONT ROUTES (observed):
- /[country]/(landing) - landing page
- /[country]/group-buying - catalog (SRCH-like)
- /[country]/group-buying/[id] - deal detail (DETL-like)
- /[country]/checkout - Medusa checkout (partial CHKO)
- /[country]/account - login + dashboard
- account/group-deals/hosted - host deals (partial DASH/MGBY/CRTE)
- account/group-deals/participations - participations (partial MYJN/MYJD)
- account/settlements - settlements (partial STLM/MSTL)
- account/payment-methods, preferences, profile, orders, addresses
"""

summary = """
================================================================================
STRUCTURED SUMMARY (for parent agent)
================================================================================

FILE 1: 아이돌공구-화면정의서.xlsx
Sheets: 가이드, 화면목록, 화면정의서, 공통규칙및미결정
Column headers:
- 가이드: 항목 | 설명 | 담당 역할
- 화면목록: 구분 | 화면ID | 화면명 | 진입 경로 | 다음 화면 | 연결 기능ID | 우선순위
- 화면정의서: 구분 | 화면ID | 화면명 | 영역 | 구성요소 | 표시 문구·데이터 | 사용자 액션 | 화면이동 | 예외·에러 처리 | 연결 기능ID | 우선순위
- 공통규칙및미결정: 분류 | 항목 | 설명 | 담당

32 screens (IDs): SPL, LGN, SGN, PAY, HOME, SRCH, ALRT, DETL, APLY, CHKO, DONE, MYJN, MYJD, RVEW, CLAM, CRTE, DASH, QFIL, SEAT, PACK, PURC, SHIP, STLM, MYPG, MPAY, MGBY, MJPT, MSTL, MTRS, MINF, MALM, MCS
Cross-cutting modules (not separate screens): TRST-01, ESCR-01/02, DEPO-01, VERI-01, LIVE-01

Key flows:
- Onboarding: SPL (splash/token) -> LGN or HOME; SGN 3-step (basic, idol select, role); PAY payment method -> HOME
- Participant: HOME/SRCH -> DETL -> APLY (shipping) -> CHKO (escrow deposit) -> DONE -> MYJN/MYJD timeline
- Host: HOME -> CRTE (open deal) -> DASH -> QFIL (fill seats) / SEAT (reject) -> PACK -> PURC (1st purchase proof) -> SHIP -> STLM (settlement)
- My Page hub MYPG: MPAY, MGBY, MJPT, MSTL, MTRS, MINF, MALM, MCS

FILE 2: 아이돌공구-기능정의서.xlsx
Sheet: 기능정의서
Columns: 구분 | 화면ID | 화면명 | 기능ID | 기능명 | 상세 설명 | 입력값 | 출력/결과 | 비고/연동 | 우선순위
~60 feature rows mapping SPL-01 through MCS-01 plus trust layer TRST, ESCR, DEPO, VERI, LIVE

Notable fields/UI:
- LGN: email, password, social (Kakao/Naver P1), forgot password
- SGN: nickname, email, password, idol chips, role participant vs host
- PAY/MPAY: bank/card/PG tokenization
- SRCH: idol/group/region filters, waitlist (ALRT), card badges (seat status, trust)
- DETL: seat list, host trust, purchase proof upload, 6-step progress timeline
- APLY: name, phone, address, split shipping, fee breakdown (seat + platform + shipping n:1)
- CHKO: escrow bank transfer until delivery confirm
- Host DASH: KPI cards, order table, deposit confirm, seat fill graph, AI price (P2)
- QFIL-02 links SRCH-04 waitlist to push matching
- ESCR: escrow hold/release; DEPO: host deposit; STLM: settlement after delivery confirm

Open decisions (from 공통규칙및미결정):
- Auto delivery confirm N days
- Host deposit amount policy
- Platform fee formula
- Shipping n:1 refund rules
- Price change after open
- Multi-seat one shipping request
- Escrow legal/PG constraints

GAP ANALYSIS vs group-buying-site storefront:
LIKELY PARTIAL / EXISTS:
- Landing (SPL/HOME marketing) -> (landing) route + landing modules
- Group deal list/detail (SRCH/DETL) -> /group-buying, /group-buying/[id], waitlist-form, join-deal-form
- Account login/register -> Medusa account @login (may not match 3-step SGN idol flow)
- Hosted deals / participations (DASH/MYJN/MGBY/MJPT) -> group-deals/hosted, participations, timelines
- Settlements (STLM/MSTL) -> account/settlements
- Payment methods (PAY/MPAY) -> payment-methods, stripe-setup-form
- Checkout -> standard Medusa checkout (may not be escrow CHKO flow)
- confirm-delivery-button -> partial MYJD delivery confirm / escrow release

LIKELY MISSING or WEAK:
- Dedicated screen IDs/routes: SPL, ALRT popup, APLY separate step, DONE, MYJD dedicated page, full host ops: CRTE wizard, DASH dashboard, QFIL, SEAT, PACK, PURC proof, SHIP tracking bulk, STLM host settlement UI
- MYPG hub with 7 sub-screens (MALM, MCS, MTRS, MINF as spec)
- Trust layer UI (TRST scores on cards), LIVE realtime, VERI fraud log UI
- Participant search filters (group/region/price), favorite idol recommendations HOME-03
- Role-based HOME (participant vs host toggle HOME-02)
- Escrow bank transfer CHKO vs card checkout
- Reviews RVEW, claims CLAM popups
- Social login LGN-02
- AI pricing DASH-05 (P2)

================================================================================
END STRUCTURED SUMMARY
================================================================================
"""

out_lines = ["IDOL GROUP-BUY SPEC EXTRACTION", ""]
for f in [screen, func]:
    if f.exists():
        extract_workbook(f, out_lines)
    else:
        out_lines.append("MISSING: " + str(f))

out_lines.append(STOREFRONT.strip())
out_lines.append(summary.strip())
OUT.write_text("\n".join(out_lines), encoding="utf-8")
print("OK", OUT.stat().st_size)
