import os
from pathlib import Path
import openpyxl

BASE = Path(r"C:\Users\uriak_wboloix\medusa\group-buying-site")
FILES = [
    BASE / "\uc544\uc774\ub3cc\uacf5\uad6c-\ud654\uba74\uc815\uc758\uc11c.xlsx",
    BASE / "\uc544\uc774\ub3cc\uacf5\uad6c-\uae30\ub2a5\uc815\uc758\uc11c.xlsx",
]
OUT = BASE / "_spec_extract.txt"
MAX_ROWS = 100

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

def extract_workbook(path, out_lines):
    out_lines.append("=" * 80)
    out_lines.append("FILE: " + path.name)
    out_lines.append("PATH: " + str(path))
    out_lines.append("=" * 80)
    if not path.exists():
        out_lines.append("ERROR: file not found")
        return
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out_lines.append("SHEET_COUNT: " + str(len(wb.sheetnames)))
    out_lines.append("SHEET_NAMES: " + repr(wb.sheetnames))
    out_lines.append("")
    for sn in wb.sheetnames:
        ws = wb[sn]
        out_lines.append("-" * 60)
        out_lines.append("SHEET: " + sn)
        out_lines.append("max_row=" + str(ws.max_row) + ", max_column=" + str(ws.max_column))
        out_lines.append("-" * 60)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= MAX_ROWS:
                out_lines.append("... truncated after " + str(MAX_ROWS) + " rows ...")
                break
            cells = [cell_str(c) for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if not any(cells):
                continue
            out_lines.append("\t".join(cells))
            row_count += 1
        out_lines.append("(rows written: " + str(row_count) + ")")
        out_lines.append("")
    wb.close()

out_lines = ["IDOL GROUP-BUY SPEC EXTRACTION", ""]
for f in FILES:
    extract_workbook(f, out_lines)
OUT.write_text("\n".join(out_lines), encoding="utf-8")
print("Wrote", OUT, OUT.stat().st_size, "bytes")
for f in FILES:
    if f.exists():
        wb = openpyxl.load_workbook(f, read_only=True)
        print(f.name, wb.sheetnames)
        wb.close()
