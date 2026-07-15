"""Extract updated idol group-buy Excel specs to _spec_extract_v2.txt."""
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
FILES = [
    BASE / "아이돌공구-기능정의서 (1).xlsx",
    BASE / "아이돌공구-화면별데이터목록.xlsx",
    BASE / "아이돌공구-화면정의서 (2).xlsx",
]
OUT = BASE / "_spec_extract_v2.txt"
MAX_ROWS = 400


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def extract_workbook(path: Path, lines: list[str]) -> None:
    lines.append("=" * 80)
    lines.append(f"FILE: {path.name}")
    if not path.exists():
        lines.append("ERROR: file not found")
        return

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    lines.append(f"SHEET_NAMES: {workbook.sheetnames!r}")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        lines.append("-" * 60)
        lines.append(f"SHEET: {sheet_name}")
        row_count = 0
        for row in worksheet.iter_rows(values_only=True):
            if row_count >= MAX_ROWS:
                lines.append(f"... truncated after {MAX_ROWS} rows ...")
                break
            cells = [cell_str(cell) for cell in row]
            while cells and cells[-1] == "":
                cells.pop()
            if not any(cells):
                continue
            lines.append("\t".join(cells))
            row_count += 1
        lines.append(f"(rows written: {row_count})")
        lines.append("")

    workbook.close()


def main() -> None:
    lines = ["UPDATED IDOL GROUP-BUY SPEC EXTRACTION", ""]
    for file_path in FILES:
        extract_workbook(file_path, lines)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
